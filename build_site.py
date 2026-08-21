# -*- coding: utf-8 -*-
"""
build_site.py - 商品期权信息汇总面板 · 静态站点构建脚本

仿照 IV_Rank 的 build_data.py 模式：把各子服务本地缓存数据转换为静态 JSON，
输出到 static/ 目录（Cloudflare Pages 输出目录），配合前端静态页实现纯静态部署。

生成内容：
  static/sub/overview/data.json        - 商品期权成交统计（最新交易日聚合）
  static/sub/oi_dist/varieties.json    - 品种分组
  static/sub/oi_dist/data/<code>.json  - 每品种最新交易日分布/最痛点
  static/sub/vol_hist/varieties.json   - 品种分组
  static/sub/vol_hist/data/<code>.json - 每品种历史波动率（默认 window=30, startYear=auto）
  static/sub/iv_rank/api/*.json        - IV 分位（复用 IV_Rank 纯函数）
  static/sub/pcr/varieties.json        - PCR 品种列表
  static/sub/pcr/data/<key>.json       - 每品种 PCR 全量
  static/sub/gex/varieties.json        - GEX 品种分组
  static/sub/gex/data/<code>.json      - 每品种最新交易日 GEX
  static/corr/                         - 相关性静态页（原样复制）
  static/site_info.json                - 各子页数据日期汇总

运行环境：需 pandas/numpy/openpyxl/requests（复用 GEX观察/.venv）。
用法：venv/Scripts/python.exe build_site.py
"""
import json
import os
import shutil
import sys
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STATIC_DIR = os.path.join(BASE_DIR, 'static')
SUB_DIR = os.path.join(STATIC_DIR, 'sub')

# 各子服务项目根目录（数据源）
P_OVERVIEW = r'D:\WorkBuddy\商品期权成交统计'
P_OIDIST = r'D:\WorkBuddy\商品期权成交持仓分布及最痛点位绘制'
P_VOLHIST = r'D:\WorkBuddy\商品期货历史波动率估计'
P_IVRANK = r'D:\WorkBuddy\IV_Rank'
P_PCR = r'D:\WorkBuddy\商品期权PCR跟踪'
P_GEX = r'D:\WorkBuddy\GEX观察'
P_CORR = r'D:\WorkBuddy\商品行情与波动率相关性'

LATEST_DATE = '20260818'  # 各子服务缓存的最新共同交易日（构建时从缓存探测，见下方）

def makedirs(p):
    os.makedirs(p, exist_ok=True)


def detect_latest_date():
    """从各子服务缓存目录探测最新共同交易日，避免硬编码过期。"""
    dates = []
    # 成交统计缓存 data/{ex}_{YYYYMMDD}.json
    d = os.path.join(P_OVERVIEW, 'data')
    if os.path.isdir(d):
        for f in os.listdir(d):
            if f.endswith('.json') and '_' in f:
                ds = f.rsplit('_', 1)[1].replace('.json', '')
                if len(ds) == 8 and ds.isdigit():
                    dates.append(ds)
    # 最痛点缓存 cache/{ex}_{YYYYMMDD}.json
    d = os.path.join(P_OIDIST, 'cache')
    if os.path.isdir(d):
        for f in os.listdir(d):
            if f.endswith('.json') and '_' in f:
                ds = f.rsplit('_', 1)[1].replace('.json', '')
                if len(ds) == 8 and ds.isdigit():
                    dates.append(ds)
    if dates:
        return max(dates)
    return LATEST_DATE


# ================= 1. overview 成交统计 =================
def build_overview(latest):
    """从成交统计本地缓存 data/*.json 聚合最新交易日数据（不联网）。"""
    out = os.path.join(SUB_DIR, 'overview')
    makedirs(out)
    cache_dir = os.path.join(P_OVERVIEW, 'data')
    if not os.path.isdir(cache_dir):
        print('[overview] 跳过：缓存目录不存在', cache_dir)
        return

    # 文件名前缀小写（dce_/czce_/...），统一转大写与交易所代码对齐
    _EX_ALIAS = {'dce': 'DCE', 'czce': 'CZCE', 'shfe': 'SHFE', 'ine': 'INE', 'gfex': 'GFEX'}
    day_files = {}  # date -> {ex: path}
    for f in os.listdir(cache_dir):
        if not f.endswith('.json') or '_' not in f:
            continue
        ex, ds = f.split('_', 1)
        ds = ds.replace('.json', '')
        if len(ds) != 8 or not ds.isdigit():
            continue
        ex = _EX_ALIAS.get(ex.lower(), ex.upper())
        day_files.setdefault(ds, {})[ex] = os.path.join(cache_dir, f)

    # 找最近 5 个交易日（至少 3 个交易所覆盖的日期，按日期降序）
    sorted_days = sorted(day_files.keys(), reverse=True)
    trading_days = []
    for ds in sorted_days:
        if len(day_files[ds]) >= 3:
            trading_days.append(ds)
        if len(trading_days) >= 5:
            break
    if not trading_days:
        print('[overview] 跳过：无足够交易日缓存')
        return

    # 聚合逻辑（与 data_aggregator.aggregate_all_exchanges 一致，只读缓存）
    exchanges = ['DCE', 'CZCE', 'SHFE', 'INE', 'GFEX']
    ex_names = {
        'DCE': 'DCE 大商所', 'CZCE': 'CZCE 郑商所', 'SHFE': 'SHFE 上期所',
        'INE': 'INE 上期能源', 'GFEX': 'GFEX 广期所',
    }
    ex_short = {'DCE': 'DCE', 'CZCE': 'CZCE', 'SHFE': 'SHFE', 'INE': 'INE', 'GFEX': 'GFEX'}

    def load_day(ex, ds):
        p = day_files.get(ds, {}).get(ex)
        if not p or not os.path.exists(p):
            return None
        try:
            with open(p, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return None

    exchanges_result = {}
    all_varieties = []
    totals = {'fut_vol': 0.0, 'opt_vol': 0.0, 'fut_oi': 0.0, 'opt_oi': 0.0}

    for ex in exchanges:
        ex_data = []
        for ds in trading_days:
            d = load_day(ex, ds)
            if d is not None:
                ex_data.append(d)
        if not ex_data:
            continue
        # 品种聚合
        var_map = {}
        for day in ex_data:
            for item in day.get('data', []):
                code = item.get('variety_code')
                if not code:
                    continue
                v = var_map.setdefault(code, {
                    'variety_code': code,
                    'variety_name': item.get('variety_name', ''),
                    'exchange': ex,
                    'exchange_name': ex_names.get(ex, ex),
                    'futures_total_volume': 0, 'options_total_volume': 0,
                    'futures_total_turnover': 0.0, 'options_total_turnover': 0.0,
                    'futures_total_open_interest': 0, 'options_total_open_interest': 0,
                    'days': 0,
                })
                v['futures_total_volume'] += item.get('futures_volume', 0)
                v['options_total_volume'] += item.get('options_volume', 0)
                v['futures_total_turnover'] += item.get('futures_turnover', 0.0)
                v['options_total_turnover'] += item.get('options_turnover', 0.0)
                v['futures_total_open_interest'] += item.get('futures_open_interest', 0)
                v['options_total_open_interest'] += item.get('options_open_interest', 0)
                v['days'] += 1

        varieties = []
        for code, v in var_map.items():
            days = v['days'] or 1
            avg_fut = v['futures_total_volume'] / days
            avg_opt = v['options_total_volume'] / days
            avg_fut_oi = v['futures_total_open_interest'] / days
            avg_opt_oi = v['options_total_open_interest'] / days
            varieties.append({
                'variety_code': code,
                'variety_name': v['variety_name'],
                'exchange': ex,
                'exchange_name': ex_names.get(ex, ex),
                'futures_avg_volume': round(avg_fut, 2),
                'options_avg_volume': round(avg_opt, 2),
                'futures_avg_turnover': round(v['futures_total_turnover'] / days, 2),
                'options_avg_turnover': round(v['options_total_turnover'] / days, 2),
                'futures_avg_open_interest': round(avg_fut_oi, 2),
                'options_avg_open_interest': round(avg_opt_oi, 2),
                'options_to_futures_ratio': round(avg_opt / avg_fut, 4) if avg_fut > 0 else 0,
                'options_to_futures_ratio_pct': round(avg_opt / avg_fut * 100, 2) if avg_fut > 0 else 0,
                'options_oi_to_futures_ratio': round(avg_opt_oi / avg_fut_oi, 4) if avg_fut_oi > 0 else 0,
                'options_oi_to_futures_ratio_pct': round(avg_opt_oi / avg_fut_oi * 100, 2) if avg_fut_oi > 0 else 0,
                'days': days,
            })
        varieties.sort(key=lambda x: x['options_avg_volume'], reverse=True)

        ex_agg = {
            'fut_vol': sum(v['futures_avg_volume'] for v in varieties),
            'opt_vol': sum(v['options_avg_volume'] for v in varieties),
            'fut_oi': sum(v['futures_avg_open_interest'] for v in varieties),
            'opt_oi': sum(v['options_avg_open_interest'] for v in varieties),
        }
        exchanges_result[ex] = {
            'exchange': ex,
            'exchange_name': ex_names.get(ex, ex),
            'short': ex_short.get(ex, ex),
            'variety_count': len(varieties),
            'total_futures_avg_volume': round(ex_agg['fut_vol'], 2),
            'total_options_avg_volume': round(ex_agg['opt_vol'], 2),
            'ratio_pct': round(ex_agg['opt_vol'] / ex_agg['fut_vol'] * 100, 2) if ex_agg['fut_vol'] > 0 else 0,
            'total_futures_avg_open_interest': round(ex_agg['fut_oi'], 2),
            'total_options_avg_open_interest': round(ex_agg['opt_oi'], 2),
            'oi_ratio_pct': round(ex_agg['opt_oi'] / ex_agg['fut_oi'] * 100, 2) if ex_agg['fut_oi'] > 0 else 0,
            'varieties': varieties,
        }
        for k in totals:
            totals[k] += ex_agg[k]
        all_varieties.extend(varieties)

    all_varieties.sort(key=lambda x: x['options_avg_volume'], reverse=True)
    all_varieties_by_oi = sorted(all_varieties, key=lambda x: x['options_avg_open_interest'], reverse=True)

    data = {
        'end_date': trading_days[0],
        'trading_days': trading_days,
        'days_count': len(trading_days),
        'exchanges': exchanges_result,
        'summary': {
            'total_futures_avg_volume': round(totals['fut_vol'], 2),
            'total_options_avg_volume': round(totals['opt_vol'], 2),
            'total_varieties': len(all_varieties),
            'overall_ratio_pct': round(totals['opt_vol'] / totals['fut_vol'] * 100, 2) if totals['fut_vol'] > 0 else 0,
            'total_futures_avg_open_interest': round(totals['fut_oi'], 2),
            'total_options_avg_open_interest': round(totals['opt_oi'], 2),
            'overall_oi_ratio_pct': round(totals['opt_oi'] / totals['fut_oi'] * 100, 2) if totals['fut_oi'] > 0 else 0,
        },
        'all_varieties': all_varieties,
        'all_varieties_by_oi': all_varieties_by_oi,
    }
    payload = {'success': True, 'data': data}
    with open(os.path.join(out, 'data.json'), 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False)
    print(f"[overview] 生成 data.json：{len(all_varieties)} 品种，{len(trading_days)} 个交易日，最新 {trading_days[0]}")


# ================= 2. oi_dist 成交持仓分布/最痛点 =================
def build_oi_dist(latest):
    """从 cache/{ex}_{date}.json 为每个品种生成最新交易日分布数据。"""
    out = os.path.join(SUB_DIR, 'oi_dist')
    data_dir = os.path.join(out, 'data')
    makedirs(data_dir)
    if P_OIDIST not in sys.path:
        sys.path.insert(0, P_OIDIST)
    try:
        from analyzer import analyze_variety
        from varieties import get_varieties_grouped
    except Exception as e:
        print('[oi_dist] 跳过：import 失败', e)
        return

    # varieties.json
    groups = get_varieties_grouped()
    with open(os.path.join(out, 'varieties.json'), 'w', encoding='utf-8') as f:
        json.dump({'success': True, 'groups': groups}, f, ensure_ascii=False)

    # 每品种数据
    ok, fail = 0, 0
    for g in groups:
        for v in g['varieties']:
            code = v['code']
            try:
                r = analyze_variety(code, latest)
                if r.get('success'):
                    with open(os.path.join(data_dir, f'{code}.json'), 'w', encoding='utf-8') as f:
                        json.dump(r, f, ensure_ascii=False)
                    ok += 1
                else:
                    fail += 1
            except Exception as e:
                fail += 1
    print(f"[oi_dist] 生成 {ok} 个品种（失败 {fail}），交易日 {latest}")


# ================= 3. vol_hist 历史波动率 =================
def build_vol_hist(latest):
    """从 .cache/{code}.csv 计算默认参数（window=30, startYear=auto）波动率。"""
    out = os.path.join(SUB_DIR, 'vol_hist')
    data_dir = os.path.join(out, 'data')
    makedirs(data_dir)
    # 用 importlib 按文件路径定向加载，避免 varieties.py 模块名跨项目冲突
    try:
        import importlib.util
        import numpy as np  # noqa: F401
        import pandas as pd

        def _load(name, path):
            spec = importlib.util.spec_from_file_location(name, path)
            mod = importlib.util.module_from_spec(spec)
            sys.modules[name] = mod
            spec.loader.exec_module(mod)
            return mod

        _vol = _load('volh_volatility', os.path.join(P_VOLHIST, 'volatility.py'))
        _var = _load('volh_varieties', os.path.join(P_VOLHIST, 'varieties.py'))
        compute_all = _vol.compute_all
        ANNUALIZATION = _vol.ANNUALIZATION
        get_varieties_grouped = _var.get_varieties_grouped
        get_variety = _var.get_variety
    except Exception as e:
        print('[vol_hist] 跳过：import 失败', e)
        return

    cache_dir = os.path.join(P_VOLHIST, '.cache')
    if not os.path.isdir(cache_dir):
        print('[vol_hist] 跳过：缓存目录不存在')
        return

    # 有缓存的品种
    codes = [f.replace('.csv', '') for f in os.listdir(cache_dir) if f.endswith('.csv')]

    # varieties.json（只列有缓存的品种，避免前端选到无数据项）
    all_groups = get_varieties_grouped()
    groups_out = []
    for g in all_groups:
        items = [it for it in g['items'] if it['code'] in codes]
        if items:
            groups_out.append({'exchange': g['exchange'], 'exchange_name': g['exchange_name'], 'items': items})
    with open(os.path.join(out, 'varieties.json'), 'w', encoding='utf-8') as f:
        json.dump(groups_out, f, ensure_ascii=False)

    ok, fail = 0, 0
    for code in codes:
        csv_path = os.path.join(cache_dir, f'{code}.csv')
        try:
            idx = pd.read_csv(csv_path, parse_dates=['date'], index_col='date')
            if idx.empty:
                fail += 1
                continue
            vol = compute_all(
                {'open': idx['open'], 'high': idx['high'], 'low': idx['low'], 'close': idx['close']},
                window=30, ann=ANNUALIZATION,
            )
            dates = [d.strftime('%Y-%m-%d') for d in idx.index]
            variety = get_variety(code)
            result = {
                'symbol': code,
                'name': variety['name'] if variety else code,
                'exchange': variety['exchange'] if variety else '',
                'category': variety['category'] if variety else '',
                'window': 30,
                'annualization': ANNUALIZATION,
                'start_year': 'auto',
                'dates': dates,
                'price': [None if (v is None or (isinstance(v, float) and np.isnan(v))) else float(v) for v in idx['close'].tolist()],
                'vol': {k: [None if (v is None or (isinstance(v, float) and np.isnan(v))) else float(v) for v in arr.tolist()] for k, arr in vol.items()},
                'points': len(dates),
                'start': dates[0] if dates else None,
                'end': dates[-1] if dates else None,
            }
            with open(os.path.join(data_dir, f'{code}.json'), 'w', encoding='utf-8') as f:
                json.dump(result, f, ensure_ascii=False)
            ok += 1
        except Exception as e:
            fail += 1
    print(f"[vol_hist] 生成 {ok} 个品种（失败 {fail}），window=30")


# ================= 4. iv_rank IV 分位（复用 IV_Rank） =================
def build_iv_rank(latest):
    """复用 IV_Rank server 纯函数生成 api/*.json（与 IV_Rank build_data.py 同源）。"""
    out = os.path.join(SUB_DIR, 'iv_rank')
    api_dir = os.path.join(out, 'api')
    makedirs(api_dir)
    if P_IVRANK not in sys.path:
        sys.path.insert(0, P_IVRANK)
    try:
        import server as ivserver
    except Exception as e:
        print('[iv_rank] 跳过：import 失败', e)
        return

    build_days = 500
    overview = ivserver.get_overview(build_days)
    with open(os.path.join(api_dir, 'overview.json'), 'w', encoding='utf-8') as f:
        json.dump({'data': overview, 'count': len(overview)}, f, ensure_ascii=False)

    for item in overview:
        code = item['code']
        detail = ivserver.get_detail(code, build_days)
        if detail:
            with open(os.path.join(api_dir, f'detail_{code}.json'), 'w', encoding='utf-8') as f:
                json.dump(detail, f, ensure_ascii=False)

    latest_d = max((it['current_date'] for it in overview), default='-')
    info = {
        'symbol_count': len(overview),
        'latest_date': latest_d,
        'excel_path': ivserver.EXCEL_PATH,
    }
    with open(os.path.join(api_dir, 'info.json'), 'w', encoding='utf-8') as f:
        json.dump(info, f, ensure_ascii=False)

    # 前端静态资源（index.html + echarts 从 IV_Rank/static 复制，改写绝对路径）
    src_html = os.path.join(P_IVRANK, 'static', 'index.html')
    dst_html = os.path.join(out, 'index.html')
    if os.path.exists(src_html):
        with open(src_html, 'r', encoding='utf-8') as f:
            html = f.read()
        # 静态模式：绝对路径 echarts 改为相对路径
        html = html.replace('src="/echarts.min.js"', 'src="./echarts.min.js"')
        html = html.replace("'/echarts.min.js'", "'./echarts.min.js'")
        with open(dst_html, 'w', encoding='utf-8') as f:
            f.write(html)
    src_ech = os.path.join(P_IVRANK, 'static', 'echarts.min.js')
    if os.path.exists(src_ech) and not os.path.exists(os.path.join(out, 'echarts.min.js')):
        shutil.copy2(src_ech, os.path.join(out, 'echarts.min.js'))
    print(f"[iv_rank] 生成 {len(overview)} 个品种 api/*.json，latest {latest_d}")


# ================= 5. pcr PCR 跟踪 =================
def build_pcr(latest):
    """从 data/tables/*.csv 生成每品种全量 PCR 数据。"""
    out = os.path.join(SUB_DIR, 'pcr')
    data_dir = os.path.join(out, 'data')
    makedirs(data_dir)
    tables_dir = os.path.join(P_PCR, 'data', 'tables')
    if not os.path.isdir(tables_dir):
        print('[pcr] 跳过：tables 目录不存在')
        return

    def read_csv(name):
        path = os.path.join(tables_dir, name)
        out_d = {}
        with open(path, 'r', encoding='utf-8', newline='') as f:
            import csv
            r = csv.reader(f)
            header = next(r)
            keys = header[1:]
            for k in keys:
                out_d[k] = []
            dates = []
            for row in r:
                dates.append(row[0])
                for k, cell in zip(keys, row[1:]):
                    out_d[k].append(None if cell == '' else float(cell))
        return dates, out_d

    with open(os.path.join(tables_dir, 'varieties.json'), 'r', encoding='utf-8') as f:
        varieties = json.load(f)

    dates, price = read_csv('price.csv')
    _, call_oi = read_csv('call_oi.csv')
    _, put_oi = read_csv('put_oi.csv')

    with open(os.path.join(out, 'varieties.json'), 'w', encoding='utf-8') as f:
        json.dump({
            'varieties': varieties,
            'date_range': {'first': dates[0] if dates else '', 'last': dates[-1] if dates else ''},
        }, f, ensure_ascii=False)

    ok = 0
    for v in varieties:
        key = v['key']
        if key not in price:
            continue
        p_arr, c_arr, put_arr = price[key], call_oi.get(key, []), put_oi.get(key, [])
        out_d, out_p, out_c, out_put, out_pcr, out_npcr = [], [], [], [], [], []
        for i, d in enumerate(dates):
            c = c_arr[i] if i < len(c_arr) and c_arr[i] is not None else 0.0
            p = put_arr[i] if i < len(put_arr) and put_arr[i] is not None else 0.0
            pr = p_arr[i] if i < len(p_arr) else None
            pcr = (p / c) if c and c > 0 else None
            denom = p + c
            npcr = ((p - c) / denom) if denom and denom > 0 else None
            out_d.append(d)
            out_p.append(pr)
            out_c.append(None if c == 0 else c)
            out_put.append(None if p == 0 else p)
            out_pcr.append(pcr)
            out_npcr.append(npcr)
        payload = {
            'key': key, 'dates': out_d, 'price': out_p,
            'call_oi': out_c, 'put_oi': out_put, 'pcr': out_pcr, 'norm_pcr': out_npcr,
        }
        # Windows 文件名不允许冒号：CZCE:AP -> CZCE_AP（前端 fetch 时同样替换）
        fname = key.replace(':', '_')
        with open(os.path.join(data_dir, f'{fname}.json'), 'w', encoding='utf-8') as f:
            json.dump(payload, f, ensure_ascii=False)
        ok += 1
    print(f"[pcr] 生成 {ok} 个品种，数据区间 {dates[0]} ~ {dates[-1]}")


# ================= 6. gex GEX 分析 =================
def build_gex(latest):
    """为每个品种生成最新交易日 GEX 数据（纯缓存模式，不联网不启动浏览器）。

    数据源：
      - 合约数据：最痛点项目 cache/{ex}_{date}.json（analyzer 归一化）
      - 标的价 S / IV：GEX观察 market_cache/（有缓存用真实值；无缓存降级 ATM 近似 + IV 0.20）
    计算：gex_core 纯函数（Black-76 反解 IV、GEX 聚合、Gamma Flip、Regime 分类）
    """
    out = os.path.join(SUB_DIR, 'gex')
    data_dir = os.path.join(out, 'data')
    makedirs(data_dir)
    for d in (P_OIDIST, P_GEX):
        if d not in sys.path:
            sys.path.insert(0, d)
    try:
        from analyzer import analyze_variety
        from varieties import get_varieties_grouped
        import gex_core
        import commodity_expiry
    except Exception as e:
        print('[gex] 跳过：import 失败', e)
        return

    # 市场缓存目录（标的价 S / IV）
    market_cache_dir = os.path.join(P_GEX, 'market_cache')

    def load_market(exchange, code):
        """只读 market_cache：有缓存返回 (underlying, options)，无缓存返回 ({}, {})。"""
        if exchange == 'DCE':
            p = os.path.join(market_cache_dir, f'dce_{code}_{latest}.json')
        else:
            p = os.path.join(market_cache_dir, f'{exchange.lower()}_{latest}.json')
        if not os.path.exists(p):
            return {}, {}
        try:
            with open(p, 'r', encoding='utf-8') as f:
                packed = json.load(f)
            underlying = {}
            options = {}
            for key, s in packed.get('underlying', {}).items():
                v, month = key.split('|')
                if v == code:
                    underlying[month] = s
            for key, info in packed.get('options', {}).items():
                rest, typ = key.rsplit('|', 1)
                mk, k = rest.rsplit('|', 1)
                v, month = mk.split('|')
                if v == code:
                    options[(month, float(k), typ)] = info
            return underlying, options
        except Exception:
            return {}, {}

    # 合约单位表（与 backend_gex.CONTRACT_UNITS 一致）
    CONTRACT_UNITS = {
        "m": 10, "c": 10, "i": 100, "pg": 20, "l": 5, "v": 5, "pp": 5, "p": 10,
        "a": 10, "b": 10, "y": 10, "eg": 10, "eb": 5, "jd": 10, "cs": 10, "lh": 16, "lg": 90,
        "cu": 5, "al": 5, "zn": 5, "pb": 5, "ni": 1, "sn": 1, "au": 1000, "ag": 15,
        "rb": 10, "ru": 10, "bu": 10, "fu": 10, "sp": 10, "ao": 20, "br": 5, "op": 20, "ad": 5,
        "sc": 1000, "nr": 10, "bc": 5,
        "si": 5, "lc": 1, "ps": 5,
        "SR": 10, "CF": 5, "TA": 5, "MA": 10, "RM": 10, "ZC": 100, "OI": 10, "PK": 5,
        "PF": 5, "PX": 5, "SH": 20, "SA": 20, "SM": 5, "SF": 5, "UR": 20, "AP": 10,
        "CJ": 5, "FG": 20, "PR": 5, "PL": 5,
    }

    def atm_strike(month_block):
        """最大持仓行权价，作为标的价格近似。"""
        best, best_oi = None, -1
        for s, co, po in zip(month_block['strikes'], month_block['call_oi'], month_block['put_oi']):
            tot = (co or 0) + (po or 0)
            if tot > best_oi:
                best_oi, best = tot, s
        return best

    def solve_month_iv(month, strikes, S, T, opt_type, oi_list, market_opts):
        """逐行权价反解 IV：Black-76 反解 -> 交易所公布IV -> None(上层用 ATM 兜底)。"""
        out = {}
        typ_key = 'C' if opt_type == 'call' else 'P'
        for s, oi in zip(strikes, oi_list):
            if not oi or oi <= 0:
                continue
            info = market_opts.get((month, float(s), typ_key))
            if not info:
                continue
            iv = gex_core.implied_vol_black76(info.get('price'), S, s, T, opt_type)
            if iv is None:
                iv = info.get('iv_pub')
            if iv and iv > 0:
                out[s] = iv
        return out

    groups = get_varieties_grouped()
    with open(os.path.join(out, 'varieties.json'), 'w', encoding='utf-8') as f:
        json.dump({'success': True, 'groups': groups}, f, ensure_ascii=False)

    from datetime import datetime as _dt
    trade_dt = _dt.strptime(latest, '%Y%m%d').date()

    ok, fail = 0, 0
    fails = []
    for g in groups:
        for v in g['varieties']:
            code = v['code']
            try:
                analysis = analyze_variety(code, latest)
                if not analysis.get('success'):
                    fail += 1
                    fails.append(code)
                    continue
                market_und, market_opts = load_market(analysis['exchange'], code)
                market_ok = bool(market_und) or bool(market_opts)

                multiplier = CONTRACT_UNITS.get(code, 10)
                out_months = []
                for mb in analysis.get('months', []):
                    strikes = mb['strikes']
                    if not strikes:
                        continue
                    month = mb['month']
                    # 标的价 S：期货结算价 > ATM 近似
                    if market_und.get(month):
                        S_use, s_src = market_und[month], 'futures'
                    else:
                        S_use, s_src = atm_strike(mb), 'atm_approx'
                    # 剩余期限 T
                    exp = commodity_expiry.expiry_date(code, analysis['exchange'], month)
                    T_days = max((exp - trade_dt).days, 1)
                    T = T_days / 365.0
                    # 逐合约 IV
                    iv_call = solve_month_iv(month, strikes, S_use, T, 'call', mb['call_oi'], market_opts)
                    iv_put = solve_month_iv(month, strikes, S_use, T, 'put', mb['put_oi'], market_opts)
                    iv_source = 'computed' if (iv_call or iv_put) else 'default'
                    merged = {}
                    merged.update(iv_put)
                    merged.update(iv_call)
                    atm_iv = merged[min(merged, key=lambda k: abs(k - S_use))] if merged else 0.20

                    def iv_for(s, side_map):
                        return side_map.get(s, atm_iv)

                    contracts = []
                    for s, oi in zip(strikes, mb['call_oi']):
                        if oi and oi > 0:
                            contracts.append({'strike': s, 'type': 'call', 'iv': iv_for(s, iv_call),
                                              'oi': oi, 'S': S_use, 'T': T, 'r': 0.0, 'multiplier': multiplier})
                    for s, oi in zip(strikes, mb['put_oi']):
                        if oi and oi > 0:
                            contracts.append({'strike': s, 'type': 'put', 'iv': iv_for(s, iv_put),
                                              'oi': oi, 'S': S_use, 'T': T, 'r': 0.0, 'multiplier': multiplier})

                    agg = gex_core.aggregate_by_strike(contracts)
                    profile = gex_core.gex_profile(contracts, S_use)
                    flip = gex_core.find_gamma_flip(profile, S_use)
                    max_wall = max(agg, key=lambda x: abs(x['gex_net'])) if agg else None
                    total_net = round(sum(x['gex_net'] for x in agg), 2)
                    total_call = round(sum(x['gex_call'] for x in agg), 2)
                    total_put = round(sum(x['gex_put'] for x in agg), 2)
                    norm = gex_core.normalized_gex(total_call, total_put, total_net)
                    regime = gex_core.classify_regime(norm)
                    out_months.append({
                        'month': month,
                        'max_pain': mb.get('max_pain'),
                        'atm_strike': atm_strike(mb),
                        'total_call_oi': mb.get('total_call_oi', 0),
                        'total_put_oi': mb.get('total_put_oi', 0),
                        'T_days': T_days,
                        'S_used': round(S_use, 4),
                        'S_source': s_src,
                        'iv_atm': round(atm_iv, 4),
                        'iv_source': iv_source,
                        'strikes': strikes,
                        'gex_call': [x['gex_call'] for x in agg],
                        'gex_put': [x['gex_put'] for x in agg],
                        'gex_net': [x['gex_net'] for x in agg],
                        'profile': profile,
                        'gamma_flip': flip,
                        'max_wall': max_wall,
                        'total_net_gex': total_net,
                        'total_call_gex': total_call,
                        'total_put_gex': total_put,
                        'normalized_gex': round(norm, 4),
                        'regime': regime,
                    })

                data_note = ('S=各月标的期货结算价；IV=期权结算价经Black-76反解(交易所公布IV/月ATM/0.20兜底)'
                             if market_ok else '行情缓存缺失，已降级：S=ATM近似、IV=0.20')
                r = {
                    'success': True,
                    'variety_code': code,
                    'variety_name': analysis['variety_name'],
                    'exchange': analysis['exchange'],
                    'exchange_name': analysis['exchange_name'],
                    'trade_date': latest,
                    'params': {'S': None, 'iv': None, 'multiplier': multiplier,
                               'T_days': None, 'market_ok': market_ok,
                               'T_note': 'T_days=None 时按交易所规则计算真实到期日(自然日)',
                               'data_note': data_note},
                    'months': out_months,
                }
                with open(os.path.join(data_dir, f'{code}.json'), 'w', encoding='utf-8') as f:
                    json.dump(r, f, ensure_ascii=False)
                ok += 1
            except Exception as e:
                fail += 1
                fails.append(f'{code}({type(e).__name__})')
    print(f"[gex] 生成 {ok} 个品种（失败 {fail}）{fails[:8]}")


# ================= 7. corr 相关性（静态复制 + xlsx 静态化） =================
def build_corr():
    """复制相关性页到 static/corr/，并把 Corr.xlsx 解析为 data.json。

    静态版不再依赖浏览器加载 xlsx（fetch 失败/公式缓存/列名尾空格等坑），
    改为构建时用 openpyxl 提取价格/IV 两个矩阵，前端直接 fetch data.json。
    """
    src = P_CORR
    dst = os.path.join(STATIC_DIR, 'corr')
    if not os.path.isdir(src):
        print('[corr] 跳过：源目录不存在')
        return
    makedirs(dst)
    for name in os.listdir(src):
        # 排除本地服务端脚本与无关文件（静态站只保留浏览器资源）
        if name.startswith('.') or name in ('start.bat', 'server.py', '.workbuddy'):
            continue
        s = os.path.join(src, name)
        d = os.path.join(dst, name)
        if os.path.isdir(s):
            # 兼容沙箱回收站限制：先删文件再删空目录，避免 rmtree 被劫持
            if os.path.exists(d):
                for root, dirs, files in os.walk(d, topdown=False):
                    for fn in files:
                        try:
                            os.remove(os.path.join(root, fn))
                        except OSError:
                            pass
                    for dn in dirs:
                        try:
                            os.rmdir(os.path.join(root, dn))
                        except OSError:
                            pass
                try:
                    os.rmdir(d)
                except OSError:
                    pass
            shutil.copytree(s, d)
        else:
            shutil.copy2(s, d)
    # 静态版前端不再需要浏览器端 SheetJS（改 fetch data.json）
    xlsx_lib = os.path.join(dst, 'libs', 'xlsx.full.min.js')
    if os.path.isfile(xlsx_lib):
        try:
            os.remove(xlsx_lib)
        except OSError:
            pass
    print('[corr] 复制到 static/corr/ 完成')
    _gen_corr_json(dst)


def _gen_corr_json(dst):
    """把 Corr.xlsx 解析为 data.json（价格/IV 两个矩阵），供静态页直接 fetch。

    关键处理：
      - 表头行定位：首列等于 Date/日期 的那一行（xlsx 前 3 行为元数据）
      - 列名统一 strip 尾空格（否则 compute.js 的 basePrice/baseIv 无法配对）
      - 日期转 'YYYY-MM-DD' 字符串，空值保留 None
      - 工作表识别：按列名含 FI.WI（价格）或 IV.WI（IV）统计
    """
    xlsx = os.path.join(dst, 'Corr.xlsx')
    if not os.path.isfile(xlsx):
        print('[corr] 跳过：无 Corr.xlsx')
        return
    try:
        import openpyxl
    except Exception as e:
        print(f'[corr] openpyxl 不可用，跳过 data.json: {e}')
        return

    def _find_header(mat):
        """定位真实表头行：首列等于 Date/日期 且列名含 FI.WI/IV.WI（跳过无列名的元数据行）。"""
        for i, r in enumerate(mat):
            if r and r[0] is not None and str(r[0]).strip().lower() in ('date', '日期'):
                cols = [str(c).strip() for c in r[1:] if c is not None]
                if any('FI.WI' in c or 'IV.WI' in c for c in cols):
                    return i
        return None

    def _sheet_kind(mat):
        """按表头列名统计识别价格/IV 工作表。"""
        hidx = _find_header(mat)
        if hidx is None:
            return None
        h = mat[hidx]
        n_p = sum(1 for c in h[1:] if c and 'FI.WI' in str(c))
        n_i = sum(1 for c in h[1:] if c and 'IV.WI' in str(c))
        return 'price' if n_p > n_i else ('iv' if n_i > n_p else None)

    def _extract(mat):
        """提取表头行 + 数据行矩阵；列名去尾空格、日期转字符串。"""
        hidx = _find_header(mat)
        if hidx is None:
            return None
        header = [('Date' if str(c).strip().lower() in ('date', '日期') else str(c).strip())
                  for c in mat[hidx]]
        out = [header]
        for r in mat[hidx + 1:]:
            if not r or r[0] is None:
                continue
            d0 = r[0]
            ds = d0.strftime('%Y-%m-%d') if isinstance(d0, datetime) else str(d0).strip()
            out.append([ds] + [None if v is None else v for v in r[1:]])
        return out

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    try:
        sheets = {}
        for name in wb.sheetnames:
            mat = []
            for row in wb[name].iter_rows(values_only=True):
                mat.append(list(row))
            sheets[name] = mat
    finally:
        wb.close()

    price_mat = iv_mat = None
    for mat in sheets.values():
        k = _sheet_kind(mat)
        if k == 'price' and price_mat is None:
            price_mat = mat
        elif k == 'iv' and iv_mat is None:
            iv_mat = mat
    if price_mat is None or iv_mat is None:
        print('[corr] 未识别价格/IV 工作表，跳过 data.json')
        return

    price = _extract(price_mat)
    iv = _extract(iv_mat)
    if not price or not iv:
        print('[corr] 矩阵提取失败，跳过 data.json')
        return
    latest = None
    try:
        dates = [r[0] for r in price[1:] if r and r[0]]
        if dates:
            latest = max(dates)
    except Exception:
        pass
    data = {
        'generated': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'latest_date': latest,
        'price': price,
        'iv': iv,
    }
    with open(os.path.join(dst, 'data.json'), 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False)
    print(f"[corr] 生成 data.json：价格 {len(price)-1} 行 x {len(price[0])} 列，"
          f"IV {len(iv)-1} 行 x {len(iv[0])} 列，最新 {latest}")


# ================= 8. site_info.json =================
def build_site_info(latest):
    info = {
        'build_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'latest_trade_date': latest,
        'sub_pages': {
            'overview': {'name': '期权市场总览', 'latest': latest},
            'oi_dist': {'name': '期权行情信息', 'latest': latest},
            'vol_hist': {'name': '历史波动率估计', 'latest': 'cache mtime'},
            'iv_rank': {'name': '隐含波动率分析', 'latest': 'api/info.json'},
            'pcr': {'name': '商品期权PCR跟踪', 'latest': 'tables'},
            'gex': {'name': '商品期权GEX分析', 'latest': latest},
            'corr': {'name': '商品行情与波动率相关性', 'latest': 'Corr.xlsx'},
        },
    }
    with open(os.path.join(STATIC_DIR, 'site_info.json'), 'w', encoding='utf-8') as f:
        json.dump(info, f, ensure_ascii=False, indent=2)
    print('[site_info] 生成 site_info.json')


# ================= 8.5 刷新各子页 index.html 写死的快照日期 =================
def refresh_html_dates(latest):
    """把各子页 index.html 中写死的日期字面量统一替换为最新交易日。

    静态版页面为单日快照（date input disabled），日期在静态化时写死，
    若构建时不同步更新会显示过期日期（如数据已到 8/21 页面仍显示 8/18）。
    用正则替换所有 YYYY-MM-DD 字面量（当前各页内仅快照日期语义）。
    """
    import re
    dash = f'{latest[:4]}-{latest[4:6]}-{latest[6:]}'
    pat = re.compile(r'\d{4}-\d{2}-\d{2}')
    changed = 0
    for name in ('overview', 'oi_dist', 'vol_hist', 'pcr', 'gex', 'iv_rank'):
        p = os.path.join(SUB_DIR, name, 'index.html')
        if not os.path.isfile(p):
            continue
        with open(p, 'r', encoding='utf-8') as f:
            s = f.read()
        s2 = pat.sub(dash, s)
        if s2 != s:
            with open(p, 'w', encoding='utf-8') as f:
                f.write(s2)
            changed += 1
            print(f'[html] {name}/index.html 日期字面量 -> {dash}')
    if not changed:
        print(f'[html] 各页日期均已为最新 {dash}，无需更新')


def main():
    latest = detect_latest_date()
    print(f'=== 商品期权信息汇总面板 静态构建 ===')
    print(f'最新共同交易日: {latest}')
    makedirs(STATIC_DIR)
    makedirs(SUB_DIR)

    build_overview(latest)
    build_oi_dist(latest)
    build_vol_hist(latest)
    build_iv_rank(latest)
    build_pcr(latest)
    build_gex(latest)
    build_corr()
    build_site_info(latest)
    refresh_html_dates(latest)
    print('=== 构建完成 ===')


if __name__ == '__main__':
    main()
